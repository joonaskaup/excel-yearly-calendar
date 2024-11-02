import pandas as pd
import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
INPUT_FILE = 'Calendar_table.xlsx'   # Your input Excel file
OUTPUT_FILE = 'yearly_calendar.xlsx' # Output Excel file
SETTINGS_FILE = 'settings.txt'       # Settings file

# Function to read settings from settings.txt
def read_settings(file_path):
    settings = {}
    current_section = None
    with open(file_path, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue  # Skip empty lines and comments
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip()
                value = value.strip()
                if key.upper() in ['TITLE', 'COLUMN_WIDTH']:
                    settings[key.upper()] = value
                    current_section = None
                elif key.upper() in ['PHASE_COLORS', 'ROW_HEIGHTS']:
                    current_section = key.upper()
                    settings[current_section] = {}
                else:
                    if current_section:
                        if current_section == 'PHASE_COLORS':
                            settings[current_section][key] = value
                        elif current_section == 'ROW_HEIGHTS':
                            settings[current_section][key.lower()] = value
                    else:
                        # Treat as a general setting
                        settings[key] = value
            else:
                # Line without ':', ignore or handle accordingly
                continue
    return settings

# Read settings
settings = read_settings(SETTINGS_FILE)

# Get settings with defaults
TITLE = settings.get('TITLE', None)
PHASE_COLORS = settings.get('PHASE_COLORS', {})
ROW_HEIGHTS = settings.get('ROW_HEIGHTS', {'normal': 20, 'special': 50})
COLUMN_WIDTH = float(settings.get('COLUMN_WIDTH', 4.5))

# Normalize phase keys in PHASE_COLORS
PHASE_COLORS = {key.strip().lower(): value for key, value in PHASE_COLORS.items()}

# Convert ROW_HEIGHTS values to float
ROW_HEIGHTS = {k.lower(): float(v) for k, v in ROW_HEIGHTS.items()}

# Read the input Excel file
df = pd.read_excel(INPUT_FILE)

# Convert 'Start' and 'End' to datetime
df['Start'] = pd.to_datetime(df['Start'], dayfirst=True, errors='coerce')
df['End'] = pd.to_datetime(df['End'], dayfirst=True, errors='coerce')

# Remove rows with invalid 'Start' dates
df = df.dropna(subset=['Start'])

# Handle missing 'End' dates by assuming single-day events
df['End'] = df['End'].fillna(df['Start'])

# Normalize the 'Phase' column
df['Phase'] = df['Phase'].astype(str).str.strip().str.lower()

# Ensure that the phases in the data match those in PHASE_COLORS
# If you changed 'Delivery' to 'Premier' in settings.txt, do the same in your Excel data

# Determine the overall date range
min_date = df['Start'].min()
max_date = df['End'].max()

# List of years in the date range
years = sorted(set(df['Start'].dt.year).union(df['End'].dt.year))

# Initialize the workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Yearly Calendar"

# Define styles
title_font = Font(bold=True, size=24)
year_header_font = Font(bold=True, color='FFFFFF', size=20)
year_header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

month_header_font = Font(bold=True, color='000000')
month_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

week_header_font = Font(bold=True, color='000000')
week_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

event_font = Font(color='000000')
event_fill = PatternFill(fill_type=None)  # No fill by default

alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=False)

thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# Set consistent column widths (48 columns per year)
for col in range(1, 49):
    ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

current_row = 1

# Add title row if TITLE is specified
if TITLE:
    title_cell = ws.cell(row=current_row, column=1, value=TITLE)
    title_cell.font = title_font
    title_cell.alignment = alignment_center
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=48)
    ws.row_dimensions[current_row].height = 30  # Adjust height as needed
    current_row += 1

for year in years:
    # Create a list of months for the year
    months = [(year, month) for month in range(1, 13)]

    # Map months to positions (0 to 11)
    month_positions = { (year, month): idx for idx, (year, month) in enumerate(months) }

    # Define a function to map a date to a week index within the year
    def date_to_week_index(date):
        if date.year != year:
            return None  # Date is not in the current year
        # Find the month index
        month_key = (date.year, date.month)
        if month_key not in month_positions:
            return None  # Date is out of range
        month_index = month_positions[month_key]
        # Assume each month has 4 weeks (0 to 3)
        day = date.day
        if day <= 7:
            week_in_month = 0
        elif day <= 14:
            week_in_month = 1
        elif day <= 21:
            week_in_month = 2
        else:
            week_in_month = 3
        week_index = month_index * 4 + week_in_month
        return week_index

    # Filter events that occur in the current year
    df_year = df[(df['Start'].dt.year <= year) & (df['End'].dt.year >= year)].copy()

    # Map events to week indices
    df_year['Start_week'] = df_year['Start'].apply(date_to_week_index)
    df_year['End_week'] = df_year['End'].apply(date_to_week_index)

    # Adjust Start_week and End_week for events that start before or end after the current year
    df_year['Start_week'] = df_year['Start_week'].fillna(0)
    df_year['End_week'] = df_year['End_week'].fillna(47)

    # Convert week indices to integers
    df_year['Start_week'] = df_year['Start_week'].astype(int)
    df_year['End_week'] = df_year['End_week'].astype(int)

    # Sort events by start week
    df_sorted = df_year.sort_values(by='Start_week')

    # Separate events into long and short events
    long_events = []
    short_events = []

    for index, event in df_sorted.iterrows():
        event_duration_weeks = event['End_week'] - event['Start_week'] + 1
        if event_duration_weeks <= 2:
            short_events.append(event)
        else:
            long_events.append(event)

    # Assign long events to rows, ensuring no overlaps
    rows = []

    for event in long_events:
        assigned = False
        for row in rows:
            conflict = False
            for e in row:
                if not (event['End_week'] < e['Start_week'] or event['Start_week'] > e['End_week']):
                    conflict = True
                    break
            if not conflict:
                row.append(event)
                assigned = True
                break
        if not assigned:
            # Create a new row for the event
            rows.append([event])

    # Create the header rows for the year
    # Year header
    year_cell = ws.cell(row=current_row, column=1, value=str(year))
    year_cell.font = year_header_font
    year_cell.fill = year_header_fill
    year_cell.alignment = alignment_center
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=48)
    # Apply borders
    for col in range(1, 49):
        cell = ws.cell(row=current_row, column=col)
        cell.border = thin_border

    # Set row height for the year header
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # Month headers
    col = 1
    for (year_month) in months:
        month_name = calendar.month_name[year_month[1]]
        month_cell = ws.cell(row=current_row, column=col, value=month_name)
        month_cell.font = month_header_font
        month_cell.fill = month_header_fill
        month_cell.alignment = alignment_center
        ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+3)
        # Apply borders
        for c in range(col, col+4):
            cell = ws.cell(row=current_row, column=c)
            cell.border = thin_border
        col += 4

    # Set row height for the month header
    ws.row_dimensions[current_row].height = 20

    current_row += 1

    # Week headers
    col = 1
    for _ in months:
        for week in range(1, 5):
            week_cell = ws.cell(row=current_row, column=col, value=f"W{week}")
            week_cell.font = week_header_font
            week_cell.fill = week_header_fill
            week_cell.alignment = alignment_center
            # Apply borders
            week_cell.border = thin_border
            col += 1

    # Set row height for the week header
    ws.row_dimensions[current_row].height = 15

    current_row += 1

    # Event rows for long events
    event_row_height = ROW_HEIGHTS.get('normal', 20)

    for row_events in rows:
        # Initialize the cells in the row with borders
        for col in range(1, 49):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_center
        # Write the events in the row
        for _, event in pd.DataFrame(row_events).iterrows():
            start_col = event['Start_week'] + 1  # Columns start from 1
            end_col = event['End_week'] + 1
            if start_col > end_col:
                start_col, end_col = end_col, start_col  # Swap if necessary
            # Ensure start_col and end_col are within 1 and 48
            start_col = max(1, min(start_col, 48))
            end_col = max(1, min(end_col, 48))
            # Merge cells for the event
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            cell = ws.cell(row=current_row, column=start_col)
            cell.value = event['Title']
            cell.font = event_font
            cell.alignment = alignment_center
            # Apply fill based on phase
            phase = event['Phase']
            phase_key = str(phase).strip().lower()
            fill_color = PHASE_COLORS.get(phase_key, None)
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            else:
                cell.fill = event_fill  # Default fill
                print(f"Warning: No color found for phase '{phase}' in event '{event['Title']}'")
            # Apply borders to all cells spanned by the event
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = alignment_center
        # Set row height
        ws.row_dimensions[current_row].height = event_row_height
        current_row += 1

    # Special row for short events
    if short_events:
        # Initialize the cells in the row with borders
        for col in range(1, 49):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_center
        # Write the short events in the last row
        for _, event in pd.DataFrame(short_events).iterrows():
            start_col = event['Start_week'] + 1  # Columns start from 1
            end_col = event['End_week'] + 1
            if start_col > end_col:
                start_col, end_col = end_col, start_col  # Swap if necessary
            start_col = max(1, min(start_col, 48))
            end_col = max(1, min(end_col, 48))
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            cell = ws.cell(row=current_row, column=start_col)
            cell.value = event['Title']
            cell.font = event_font
            cell.alignment = alignment_center
            phase = event['Phase']
            phase_key = str(phase).strip().lower()
            fill_color = PHASE_COLORS.get(phase_key, None)
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            else:
                cell.fill = event_fill
                print(f"Warning: No color found for phase '{phase}' in event '{event['Title']}'")
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = alignment_center
        special_row_height = ROW_HEIGHTS.get('special', 50)
        ws.row_dimensions[current_row].height = special_row_height
        current_row += 1

    # Add an empty row between years for spacing
    current_row += 1

# Add the legend below the calendar after all years
legend_start_row = current_row + 1  # Leave a blank row
legend_column = 1  # Start from the first column

ws.cell(row=legend_start_row, column=legend_column, value="Legend:")
ws.cell(row=legend_start_row, column=legend_column).font = Font(bold=True)
ws.row_dimensions[legend_start_row].height = 20

legend_row = legend_start_row + 1

for phase, color in PHASE_COLORS.items():
    # Create a cell with the color fill
    color_cell = ws.cell(row=legend_row, column=legend_column)
    color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    color_cell.border = thin_border
    color_cell.alignment = alignment_center
    # Adjust the column width for the legend to accommodate phase names
    ws.column_dimensions[get_column_letter(legend_column)].width = COLUMN_WIDTH # Adjust as needed

    # Write the phase name next to the color cell
    phase_cell = ws.cell(row=legend_row, column=legend_column + 1, value=phase.title())
    phase_cell.alignment = alignment_left
    phase_cell.border = thin_border
    # Merge cells to prevent text from spilling into adjacent cells
    ws.merge_cells(start_row=legend_row, start_column=legend_column + 1, end_row=legend_row, end_column=legend_column + 5)

    legend_row += 1

# Save the workbook
wb.save(OUTPUT_FILE)

print(f"Yearly calendar has been successfully created and saved to {OUTPUT_FILE}")